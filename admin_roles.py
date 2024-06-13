import pandas as pd
import openpyxl
import os
import cv2
import streamlit as st
from PIL import Image
import shutil
from image_ret_proc import update_face_encodings

users_path = 'users.xlsx'
images_path = 'images/' 

def show_users():
    users = pd.read_excel(users_path)
    return users[users['username'] != 'admin'][['username', 'role']]
    

def add_user(username, password, role='user'):
    wb = openpyxl.load_workbook(users_path)
    users_sheet = wb.active
    new_user = [username, password, role]
    users_sheet.append(new_user)
    wb.save(users_path)
    wb.close()

def remove_user(username):
    wb = openpyxl.load_workbook(users_path)
    users_sheet = wb.active
    
    for i in range(1, users_sheet.max_row +1):
        if users_sheet.cell(row=i, column=1).value == username:
            #print(i)
            delete_row = i
            break
        
    users_sheet.delete_rows(delete_row, 1)
    wb.save(users_path)
    wb.close()
    

def capture_image(name_person):
    cam = cv2.VideoCapture(0)
    cv2.namedWindow("Capture image")
    img_counter = 0

    print('Press the SPACE key to shot a photo') 
    print('Press ESC to finish the capture')

    while True:
        ret, frame = cam.read()
        
        if not ret:
            print("failed to grab frame")
            break
        cv2.imshow("test", frame)

        k = cv2.waitKey(1)
        if k%256 == 27:
            # ESC pressed
            print("Escape hit, closing...")
            break
        elif k%256 == 32:
            # SPACE pressed
            img_name = "images/{}/{}.jpg".format(name_person, name_person + '_' + img_counter)
            cv2.imwrite(img_name, frame)
            #print("{} written!".format(img_name))
            img_counter += 1

    cam.release()
    cv2.destroyAllWindows()
    if img_counter > 0:
        print('Images saved!')


def add_person(name):
    dir_name = name.replace(' ', '-').lower()
    new_path = images_path + dir_name
    
    if not os.path.exists(new_path):
        os.makedirs(new_path)
        st.success("Added new person")
    else:
        st.warning("This person exists!")
        
def remove_person(folder_path):
    shutil.rmtree(images_path + folder_path)
    
def all_persons():
    return os.listdir(images_path) 
    #for f in os.listdir(images_path):
     #   st.write(f.replace('-', ' ').upper)
        
        
def add_images(uploaded_files, name):
    for i in range(0, len(uploaded_files)):
        
        bytes_data = uploaded_files[i].read()  # read the content of the file in binary
        
        with open(os.path.join(images_path + name + '/', uploaded_files[i].name), "wb") as f:
           f.write(bytes_data)  
           
           
    
    #update_face_encodings()
           
    st.success('Images added')
        #image = Image.open(uploaded_files[i])
        #st.image(image, width=300)
        
        #st.write("filename:", uploaded_file.name)
        #st.write(bytes_data)
    
    #return True
    

def person_images(name):
    person_imgs_path = images_path + name + '/'
    
    return [person_imgs_path + f for f in os.listdir(person_imgs_path)]

def show_image(img_path):
    image = Image.open(img_path)
    st.image(image, width=300)
    st.write("filename:", img_path)
    
    
    #cnt = 0
    #for f in os.listdir(person_imgs_path):
    #    image = Image.open(f)
    #    st.image(image, width=300)
    #    st.write("filename:", f)
    #    rem_img_btn = st.button('Remove image', key=f'rem_img_btn_{cnt}')
    #    if rem_img_btn:
    #        remove_img(person_imgs_path + f)
            
    #    cnt += 1
            
def remove_img(img):
    
    path_to_img = img
    try:
        os.remove(path_to_img)
        return True
    except IndexError:
        pass
    
    
        
    
    
    

    